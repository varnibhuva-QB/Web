"""
Export leads to CSV, Excel, or JSON.
"""
import io
import json
import csv
from datetime import datetime
from typing import List, Tuple


COLUMNS = [
    "business_name", "company_name", "phone", "email",
    "address", "website", "rating", "category",
    "product", "source", "keyword", "location", "scraped_at"
]


class ExportHandler:

    def export(self, leads: List[dict], fmt: str) -> Tuple[str, bytes, str]:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if fmt == "csv":
            return self._to_csv(leads, ts)
        elif fmt == "excel":
            return self._to_excel(leads, ts)
        elif fmt == "json":
            return self._to_json(leads, ts)
        else:
            raise ValueError(f"Unsupported format: {fmt}")

    def _to_csv(self, leads, ts):
        buf = io.StringIO()
        cols = self._detect_columns(leads)
        writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(leads)
        content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        return f"leads_{ts}.csv", content, "text/csv"

    def _to_excel(self, leads, ts):
        try:
            import pandas as pd
        except ImportError:
            raise RuntimeError("pandas not installed. Run: pip install pandas openpyxl")

        df = pd.DataFrame(leads)
        cols = [c for c in COLUMNS if c in df.columns]
        extra = [c for c in df.columns if c not in COLUMNS and c != "id"]
        df = df[cols + extra]

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Leads", index=False)
            # Style header row
            ws = writer.sheets["Leads"]
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            # Auto column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        return f"leads_{ts}.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _to_json(self, leads, ts):
        content = json.dumps(leads, indent=2, default=str).encode("utf-8")
        return f"leads_{ts}.json", content, "application/json"

    def _detect_columns(self, leads) -> list:
        seen = set()
        ordered = []
        for col in COLUMNS:
            if any(col in l for l in leads):
                ordered.append(col)
                seen.add(col)
        for lead in leads:
            for k in lead:
                if k not in seen and k != "id":
                    ordered.append(k)
                    seen.add(k)
        return ordered
